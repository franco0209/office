import time
from openpyxl import load_workbook
from main.models import Empresa, Cuenta, Agencia, Inspector
from json import JSONDecodeError, loads, dumps
import win32com.client
import os
import re
import customtkinter as ctk
from tkinter import filedialog
import unicodedata

def deserializar_referencias(cuenta: Cuenta) -> list:

    try:
        return loads(cuenta.referencias or "[]")
    except JSONDecodeError:
        return []

def serializar_referencias(referencias: list) -> str:

    return dumps(referencias, ensure_ascii=False)

def searchEmpresa(ppto):
    ppto = ppto.upper()
    pptoFilter = ppto.split()[0]
    empresa = Empresa.objects.filter(key = pptoFilter).first()
    return empresa

def searchAccountByName(addressee, cuentas):
    account = None
    namesClient = cuentas.values_list('nombre', flat=True)
    for name in namesClient:
        if name in addressee:
            account = cuentas.filter(nombre = name).first()
            return account
    return account

def searchAccountByKey(addresseeFilter, cuentas):
    account = None
    keysClient = cuentas.values_list('key', flat=True)
    for key in keysClient:
        if key in addresseeFilter:
            account = cuentas.get(key = key)
            return account
    return account

def searchAccountByReferences(addresseeFilter, cuentas):
    account = None
    for cuenta in cuentas:
        referencias = deserializar_referencias(cuenta)
        for referencia in referencias:
            if referencia in addresseeFilter:
                account = cuentas.get(nombre = cuenta.nombre)
                return account
    return account

def searchCuenta(addressee, client):
    addressee = addressee.upper()
    cuentas = Cuenta.objects.filter(empresa = client)
    cuenta = None
    cuenta = searchAccountByName(addressee, cuentas)
    if cuenta is None:
        addresseeFilter = addressee.split()
        cuenta = searchAccountByKey(addresseeFilter, cuentas)
        if cuenta is None:
            cuenta = searchAccountByReferences(addresseeFilter, cuentas)
    return cuenta

def searchAgencia(addressee, cuenta):
    addressee = addressee.upper()
    agencias = Agencia.objects.filter(cuenta = cuenta)
    agencia = None
    agencia = searchAccountByName(addressee, agencias)
    if agencia is None:
        addresseeFilter = addressee.split()
        agencia = searchAccountByKey(addresseeFilter, agencias)
        if agencia is None:
            agencia = searchAccountByReferences(addresseeFilter, agencias)
    return agencia

def searchInspectorByName(filterAddressee, inspectores):

    if isinstance(filterAddressee, str):
        search_terms = [filterAddressee.strip().lower()]
    else:
        search_terms = [term.strip().lower() for term in filterAddressee]
    
    names = inspectores.values_list('nombre', flat=True)
    normalized_names = {name.lower().strip(): name for name in names}
    
    matching_names = []
    for term in search_terms:
        if term in normalized_names:
            matching_names.append(normalized_names[term])
    
    if matching_names:
        return list(inspectores.filter(nombre__in=matching_names))
    else:
        for term in search_terms:
            for name in normalized_names.keys():
                if name in term:
                    matching_names.append(normalized_names[name])
    
    if matching_names:
        return list(inspectores.filter(nombre__in=matching_names))
    
    return []

def searchInspectorByKey(filterAddressee, inspectores):
    if isinstance(filterAddressee, str):
        search_terms = [filterAddressee.strip().lower()]
    else:
        search_terms = [term.strip().lower() for term in filterAddressee]
    
    names = inspectores.values_list('nombreClave', flat=True)
    normalized_names = {name.lower().strip(): name for name in names}
    
    matching_names = []
    for term in search_terms:
        if term in normalized_names:
            matching_names.append(normalized_names[term])
    
    if matching_names:
        return list(inspectores.filter(nombreClave__in=matching_names))
    else:
        for term in search_terms:
            for name in normalized_names.keys():
                if name in term:
                    matching_names.append(normalized_names[name])
    if matching_names:
        return list(inspectores.filter(nombreClave__in=matching_names))
    return []

def normalize_text(text):

    if not text:
        return text
    
    normalized = unicodedata.normalize('NFKD', str(text))
    return ''.join(
        c for c in normalized 
        if not unicodedata.combining(c)
    )

def searchInspector(addressee, empresa):
    inspectores = Inspector.objects.filter(empresa=empresa)
    addressee = normalize_text(addressee)
    
    if isinstance(addressee, str) and '-' in addressee:
        parts = [part.strip() for part in addressee.split('-') if part.strip()]
        return searchInspectorByName(parts, inspectores)
    
    elif isinstance(addressee, str) and '/' in addressee:
        parts = [part.strip() for part in addressee.split('/') if part.strip()]
        return searchInspectorByName(parts, inspectores)
    
    inspectors = searchInspectorByName(addressee, inspectores)

    if inspectors == []:
        inspectors = searchInspectorByKey(addressee, inspectores)
    return inspectors

def extractMonto(sheet, columna, filaActual, paso):
        while True:
            celda = sheet[f'{columna}{filaActual}']
            if celda.value is None:
                celdaSiguente = sheet[f'{columna}{filaActual + 1}']
                if celdaSiguente.value is None:
                    break
            filaActual += paso

        for fila in range(filaActual - 1, 0, -1):
            celda = sheet[f'{columna}{fila}']
            if celda.value is not None:
                try:
                    value = float(celda.value)
                    return value
                except:
                    continue
        return None
    
def leerDatosExcel(ruta):
    wb = load_workbook(ruta, data_only=True, read_only=True)
    sheet = wb.active

    datos = {
        "agenciaDestinatarios": sheet['D17'].value,
        "inspectoresDestinatarios": sheet['D18'].value,
        "referencia": sheet['D19'].value,
        "ppto": sheet['J13'].value,
        "monto": extractMonto(sheet, 'J', 50, 10)
    }
    return datos

def leerDatosTottus(ruta, nameCotizacion):
    wb = load_workbook(ruta, data_only=True, read_only=True)
    sheet = wb.active
    
    match = re.search(r'TT 00\d{2}', nameCotizacion)
    ppto = match.group(0) if match else nameCotizacion.split(".")[0]
    
    datos = {
        "agenciaDestinatarios": nameCotizacion,
        "inspectoresDestinatarios": sheet['C9'].value,
        "referencia": sheet['C10'].value,
        "ppto": ppto,
        "monto": extractMonto(sheet, 'I', 50, 10)
    }
    return datos
    
   

def obtenerCuenta(destinatarios, nombre_cotizacion, empresa):
    cuenta = searchCuenta(destinatarios, empresa)
    if cuenta is None:
        cuenta = searchCuenta(nombre_cotizacion, empresa)
    return cuenta

def obtenerAgencia(destinatarios, nombre_cotizacion, cuenta):
    agencia = searchAgencia(destinatarios, cuenta)
    if agencia is None:
        agencia = searchAgencia(nombre_cotizacion, cuenta)
    return agencia

def generarSaludo(inspectores):
    if not inspectores:
        return "Estimado Sr."
    
    if len(inspectores) > 1:
        nombres_formales = []
        for inspector in inspectores:
            if inspector.isFemale:
                nombres_formales.append(f"Srta. {inspector.nombreClave}")
            else:
                nombres_formales.append(f"Sr. {inspector.nombreClave}")
        return "Estimados " + ", ".join(nombres_formales) + ","
    
    inspector = inspectores[0]
    if inspector.isFemale:
        return f"Estimada Srta. {inspector.nombreClave}, "
    else:
        return f"Estimado Sr. {inspector.nombreClave}, "
    
def generarInformacion(codigo_servicio, monto, estado, tipo_ticket):
    mensaje = "Se solicita " if estado == "ATENDIDO" else "Para proceder con el inicio de los trabajos se requiere "

    if "Sin" in codigo_servicio:
        mensaje += f"el envío del código del servicio ({tipo_ticket}), "

    mensaje += "la aprobación del servicio por correo y la orden de compra."

    if monto < (700 / 1.18) and tipo_ticket == "TICKET":
        mensaje += " Caso contrario indicar si el pago es por planilla recurrente."

    return mensaje

def construirTabla(ppto, monto, empresa, cuenta, estado, codigo_servicio, agencia, referencia):
    return {
        "ppto": ppto,
        "costo": f"{monto} Sin IGV",
        "cliente": empresa.nombre,
        "cuenta": cuenta.nombre,
        "estado": estado,
        "codigo Del Servicio": codigo_servicio,
        "zona": agencia.ciudad,
        "agencia": agencia.nombre,
        "descripcion Del Servicio": referencia,
    }
    
def construirContexto(inspectores, saludo, informacion):
    correos = [inspector.correo for inspector in inspectores]
    return {
        "correos": correos,
        "saludo": saludo,
        "informacion": informacion,
    }
def extraerCodigoServicio(nombre_folder, tipo_ticket):
    if tipo_ticket == "OT":
        regex = r"OT\s?\d{4}"
        match = re.findall(regex, nombre_folder)
        if match:
            return match[0]    
    else: 
        regex = r"[A-Z]{2}\d{6}\.\d{2}"
        match = re.findall(regex, nombre_folder)
        if match:
            return match[0]
        
    return f"Sin {tipo_ticket}"

    
def extractData(cotizacion, nameCotizacion, nameFolder, entregables):
    datos_excel = leerDatosExcel(cotizacion)
    if not datos_excel['agenciaDestinatarios']:
        datos_excel = leerDatosTottus(cotizacion, nameCotizacion)
        
    empresa = searchEmpresa(datos_excel['ppto'])
    cuenta = obtenerCuenta(datos_excel['agenciaDestinatarios'], nameCotizacion, empresa)
    agencia = obtenerAgencia(datos_excel['agenciaDestinatarios'], nameCotizacion, cuenta)
    inspectores = searchInspector(datos_excel['inspectoresDestinatarios'], empresa)

    saludo = generarSaludo(inspectores)
    estado = "ATENDIDO" if entregables else "NUEVO"
    codigo = extraerCodigoServicio(nameFolder, empresa.tipoTicket)
    informacion = generarInformacion(codigo, datos_excel['monto'], estado, empresa.tipoTicket)

    tabla = construirTabla(
        ppto=datos_excel['ppto'],
        monto=datos_excel['monto'],
        empresa=empresa,
        cuenta=cuenta,
        estado=estado,
        codigo_servicio=codigo,
        agencia=agencia,
        referencia=datos_excel['referencia']
    )

    contexto = construirContexto(
        inspectores=inspectores,
        saludo=saludo,
        informacion=informacion
    )

    return tabla, contexto


def iniciarOutlook():
    outlook = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")
    inbox = namespace.GetDefaultFolder(6)
    return outlook, inbox

from datetime import datetime, timedelta

def buscarCorreoOriginal(subjectToReply, inbox, dias_busqueda=10):
    fecha_limite = (datetime.now() - timedelta(days=dias_busqueda)).strftime('%m/%d/%Y %H:%M %p')
    filtro = f"[ReceivedTime] >= '{fecha_limite}'"

    try:
        items = inbox.Items.Restrict(filtro)
        items.Sort("[ReceivedTime]", True)
        
        subjetLower = subjectToReply.lower()
        for item in items:
            if subjetLower in item.Subject.lower():
                return item
    except Exception as e:
        print(f"Error en búsqueda restringida: {e}")
    
    return None

def agregarCuentaRemitente(mail, remitente):
    for account in mail.Application.Session.Accounts:
        if account.DisplayName == remitente:
            mail._oleobj_.Invoke(*(64209, 0, 8, 0, account))
            break

def agregarDestinatarios(mail, correos):
    try:
        mail.To = "; ".join(correos)
    except Exception:
        mail.To = ""

    # CC fijo
    mail.CC = (
        "<josesotom@electrototalsecurity.com>; "
        "'Analista Operaciones Electrototal' <analista.operaciones@electrototalsecurity.com>; "
        "<asistente.administrativo@electrototalsecurity.com>; "
        "<asistente.operaciones@electrototalsecurity.com>"
    )

def construirTablaHTML(tabla):
    headers = "".join(f"<th style='text-align: center; padding: 4px;'>{key.upper()}</th>" for key in tabla)
    values = "".join(f"<td style='text-align: center; padding: 4px;'>{val}</td>" for val in tabla.values())
    return f"""
        <table border='1' style='border-collapse: collapse; font-family: Calibri; color: #2F5597; font-size:9pt;'>
            <tr style='background-color: #2F5597; font-family: Arial; color: white; font-size: 10pt;'>{headers}</tr>
            <tr>{values}</tr>
        </table>
    """

def construirFirmaHTML(cid):
    return f"""<table style='border-collapse:collapse;'>
        <tr>
            <td><img src='cid:{cid}' width='274' height='97' /></td>
            <td>
                <p style='color:#6699FF; font-family:Arial; margin:0;'><b>José Luis Soto Marin</b><br/>
                Jefe de Operaciones<br/>
                999 999 944<br/>
                <a href='mailto:josesotom@electrototalsecurity.com'>josesotom@electrototalsecurity.com</a><br/>
                <a href='mailto:electrototals@gmail.com'>electrototals@gmail.com</a></p>
            </td>
        </tr>
    </table>"""

def construirCuerpoCorreo(contexto, tabla_html, firma_html, observaciones):
    
    return f"""
    <html>
        <body style='font-family: Calibri; color: #2F5597; font-size: 12pt;'>
            <p>{contexto["saludo"]}</p> 
            <p>Le enviamos nuestra cotización {observaciones} en atención del siguiente servicio:</p> 
            {tabla_html}
            <p>{contexto["informacion"]}</p> 
            <p>Quedamos atentos a sus comentarios.</p>
            <p>Saludos cordiales.</p>
            {firma_html}
        </body>
    </html>
    """

def adjuntarFirma(mail, cid):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    firma_path = os.path.join(script_dir, "firma.png")
    attachment = mail.Attachments.Add(firma_path)
    attachment.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", cid)
    
def añadirObservaciones(ppto, tieneEntregables):
    
    acumulador = 0
    observaciones = ""
    if "_" in ppto:
        acumulador += 1
    if tieneEntregables:
        acumulador +=2
        
    if acumulador == 1:
        observaciones = "actualizada"
    elif acumulador == 2:
        observaciones = "y los entregables"
    elif acumulador == 3:
        observaciones = "actualizada y los entregables"
        
    return observaciones
        

def sendEmail(tabla, contexto, adjuntos, subjectToReply):
    outlook, inbox = iniciarOutlook()
    original = None
    if subjectToReply != "":
        original = buscarCorreoOriginal(subjectToReply, inbox)
    mail = original.Reply() if original else outlook.CreateItem(0)

    agregarCuentaRemitente(mail, "cotizaciones@electrototalsecurity.com")
    agregarDestinatarios(mail, contexto["correos"])

    codigo = tabla["ppto"]
    
    observaciones = añadirObservaciones(codigo, tabla["estado"] == "ATENDIDO")
    
    if "sin" not in tabla["codigo Del Servicio"].lower():
        codigo = f"{tabla['codigo Del Servicio']}_{codigo}"
    mail.Subject = f"COTIZACIÓN {codigo} || {tabla['descripcion Del Servicio']} || {tabla['cuenta']} {tabla['agencia']}"

    cid = "firma123"
    adjuntarFirma(mail, cid)
    tabla_html = construirTablaHTML(tabla)
    firma_html = construirFirmaHTML(cid)
    cuerpo_html = construirCuerpoCorreo(contexto, tabla_html, firma_html, observaciones)
    mail.HTMLBody = cuerpo_html + mail.HTMLBody

    for adj in adjuntos:
        try:
            mail.Attachments.Add(adj)
        except Exception as e:
            print(f"Error al adjuntar {adj}: {e}")

    mail.Display()
    
def obtenerArchivosDeCotizacion(ruta_folder):
    archivos_adjuntos = []
    path_cotizacion_excel = None
    nombre_cotizacion_excel = None
    total_size = 0

    if not os.path.exists(ruta_folder):
        print("❌ La ruta no existe")
        return None, None, [], False

    for item in os.listdir(ruta_folder):
        path = os.path.join(ruta_folder, item)
        if not os.path.isfile(path):
            continue

        nombre = item.lower()
        size = os.path.getsize(path)

        if (nombre.endswith(".xlsx") and 
            "cotiza" in nombre):
            path_cotizacion_excel = path
            nombre_cotizacion_excel = item
            continue

        if nombre.endswith((".pdf", ".docx", ".pptx")):
            if ("informe" in nombre or 
                "acta" in nombre or 
                "cotiza" in nombre):
                total_size += size
                archivos_adjuntos.append(path)                

    if total_size > 14 * 1024 * 1024:
        archivos_adjuntos = [f for f in archivos_adjuntos 
                           if not f.lower().endswith((".docx", ".pptx"))]

    return path_cotizacion_excel, nombre_cotizacion_excel, archivos_adjuntos, len(archivos_adjuntos) > 1

def procesarYEnviarCorreo(path_coti, name_coti, archivos, entregables, name_folder, subject):
    if not path_coti:
        print("⚠️ No se encontró archivo de cotización")
        return

    tabla, contexto = extractData(path_coti, name_coti, name_folder, entregables)
    
    
    sendEmail(tabla, contexto, archivos, subjectToReply=subject)

def procesarCarpetaCotizacion(ruta, subject_to_find):
    name_folder = os.path.basename(ruta)
    path_coti, name_coti, archivos, entregables = obtenerArchivosDeCotizacion(ruta)
    
    procesarYEnviarCorreo(path_coti, name_coti, archivos, entregables, name_folder, subject_to_find)
    
def run_gui():
    def seleccionar_carpeta():
        ruta = filedialog.askdirectory()
        if ruta:
            folder_var.set(ruta)

    def ejecutar_proceso():
        ruta = folder_var.get()
        asunto = asunto_entry.get().strip()

        if not ruta or not os.path.exists(ruta):
            resultado_var.set("⚠️ Carpeta no válida.")
            return

        inicio = time.perf_counter()
        procesarCarpetaCotizacion(ruta, subject_to_find=asunto)
        fin = time.perf_counter()
        duracion = fin - inicio
        resultado_var.set(f"✅ Listo en {duracion:.2f} segundos.")

    ctk.set_appearance_mode("system")
    ctk.set_default_color_theme("blue")

    app = ctk.CTk()
    app.title("Responder Cotización")
    app.geometry("600x250")

    folder_frame = ctk.CTkFrame(app)
    folder_frame.pack(pady=10, padx=20, fill="x")

    folder_var = ctk.StringVar()
    ctk.CTkLabel(folder_frame, text="Carpeta:").pack(side="left", padx=5)
    ctk.CTkEntry(folder_frame, textvariable=folder_var, width=400).pack(side="left", padx=5)
    ctk.CTkButton(folder_frame, text="Seleccionar", command=seleccionar_carpeta).pack(side="left")

    ctk.CTkLabel(app, text="Asunto del correo a responder:").pack(pady=(10, 0))
    asunto_entry = ctk.CTkEntry(app, width=550)
    asunto_entry.pack(pady=5)

    ctk.CTkButton(app, text="Procesar y Enviar", command=ejecutar_proceso).pack(pady=10)

    resultado_var = ctk.StringVar()
    resultado_label = ctk.CTkLabel(app, textvariable=resultado_var, text_color="gray")
    resultado_label.pack(pady=5)

    app.mainloop()
    
def run():
    run_gui()