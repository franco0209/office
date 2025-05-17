from openpyxl import load_workbook
from openai import OpenAI

def connectApi(datos):
    apiKey = "sk-918a3a5bbc1b4d4e97032576083b2d04"
    client = OpenAI(api_key=apiKey, base_url="https://api.deepseek.com")
    response = client.chat.completions.create(
    model="deepseek-chat",
    messages=[
        {
            "role": "system",
            "content": "Eres un asistente que genera un título de servicio y alcances técnicos para informes de mantenimiento. Devuelve ÚNICAMENTE un array JSON de 2 strings: [servicio, alcances]. Servicio debe ser un título conciso (ej: 'mantenimiento de [equipo] Nro. [X]'). Alcances debe ser una lista de actividades en infinitivo separadas por comas, sin viñetas."
        },
        {
            "role": "user",
            "content": f"Extrae título de servicio y alcances de: '{datos}'"
        }
    ],
    stream=False
)
    print(response.choices[0].message.content)

def leer_excel_sin_espacios(ruta_archivo):
    wb = load_workbook(ruta_archivo, read_only=True, data_only=True)
    sheet = wb.active
    
    contenido = []
    
    for row in sheet.iter_rows(values_only=True):
        # Verificar si la fila tiene al menos una celda con contenido
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            # Convertir las celdas a strings, manejando valores None
            fila = [str(cell).strip() if cell is not None else "" for cell in row]
            contenido.append("\t".join(fila))
    
    return "\n".join(contenido)

# Uso:
def run():
    datos = leer_excel_sin_espacios(r"C:\Users\casa\Documents\Projects\dbases\orm\COTIZACION N° TGS 0209 - RETIRO DE BOMBA SUMERGILE - IBK CAYMA.xlsx")
    connectApi(datos)

