import os
from openpyxl import load_workbook
import re
from typing import Dict, Optional, Union
import unicodedata
from dataclasses import dataclass
import logging

# Configuración básica de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class ExcelData:
    empresa: str
    cuenta: str
    agencia: str
    inspectores: str
    referencia: str
    presupuesto: str
    monto: float
    tipo_ticket: str
    contenido_completo: str

def normalize_text(text: str) -> str:
    """Normaliza texto para búsquedas consistentes"""
    if not text:
        return ""
    
    normalized = unicodedata.normalize('NFKD', str(text))
    return ''.join(c for c in normalized if not unicodedata.combining(c)).strip().upper()

def extract_monto(sheet, columna: str, fila_inicial: int, paso: int) -> Optional[float]:
    """
    Busca el monto en una columna, comenzando desde fila_inicial y subiendo.
    """
    fila_actual = fila_inicial
    last_value = None
    
    while fila_actual > 0:
        celda = sheet[f'{columna}{fila_actual}']
        if celda.value is not None:
            try:
                value = float(celda.value)
                last_value = value
            except (ValueError, TypeError):
                pass
        
        # Si encontramos un valor y la siguiente celda está vacía, terminamos
        if last_value is not None and sheet[f'{columna}{fila_actual + 1}'].value is None:
            return last_value
            
        fila_actual -= paso
    
    return last_value

def extract_full_content(sheet) -> str:
    """Extrae todo el contenido del Excel como texto estructurado"""
    content = []
    
    for row in sheet.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
            content.append("\t".join(clean_row))
    
    return "\n".join(content)

def extract_ticket_code(nombre_archivo: str, tipo_ticket: str) -> str:
    """Extrae el código de servicio basado en el tipo de ticket"""
    if tipo_ticket == "OT":
        match = re.search(r"OT\s?\d{4}", nombre_archivo)
    else:
        match = re.search(r"[A-Z]{2}\d{6}\.\d{2}", nombre_archivo)
    
    return match.group(0) if match else f"Sin {tipo_ticket}"

def determine_ticket_type(nombre_archivo: str) -> str:
    """Determina el tipo de ticket basado en patrones en el nombre del archivo"""
    if re.search(r"OT\s?\d{4}", nombre_archivo):
        return "OT"
    return "TICKET"

def extract_excel_data(file_path: str, file_name: str) -> Dict[str, Union[str, float]]:
    """
    Extrae datos estructurados de un archivo Excel de cotización.
    
    Args:
        file_path: Ruta completa al archivo Excel
        file_name: Nombre del archivo (para extraer metadatos)
    
    Returns:
        Diccionario con todos los datos relevantes estructurados
    """
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        sheet = wb.active
        
        # Datos básicos
        datos = {
            "agencia": sheet['D17'].value,
            "inspectores": sheet['D18'].value,
            "referencia": sheet['D19'].value,
            "presupuesto": sheet['J13'].value,
            "monto": extract_monto(sheet, 'J', 50, 10),
            "tipo_ticket": determine_ticket_type(file_name),
            "contenido_completo": extract_full_content(sheet)
        }
        
        # Normalización de campos de texto
        for field in ["agencia", "inspectores", "referencia", "presupuesto"]:
            if datos[field]:
                datos[field] = normalize_text(datos[field])
        
        # Caso especial para Tottus (puede extenderse para otros clientes)
        if not datos["agencia"] or "TOTTUS" in file_name.upper():
            datos.update({
                "agencia": file_name,
                "inspectores": sheet['C9'].value,
                "referencia": sheet['C10'].value,
                "monto": extract_monto(sheet, 'I', 50, 10)
            })
            
            match = re.search(r'TT 00\d{2}', file_name)
            if match:
                datos["presupuesto"] = match.group(0)
        
        # Extraer código de servicio
        datos["codigo_servicio"] = extract_ticket_code(file_name, datos["tipo_ticket"])
        
        # Identificar empresa (primer palabra del presupuesto)
        if datos["presupuesto"]:
            datos["empresa"] = datos["presupuesto"].split()[0]
        
        return datos
        
    except Exception as e:
        logger.error(f"Error al procesar el archivo {file_path}: {str(e)}")
        raise

def find_excel_file(folder_path: str) -> Optional[str]:
    """Busca el archivo Excel de cotización en una carpeta"""
    for file in os.listdir(folder_path):
        if file.lower().endswith('.xlsx') and 'cotiza' in file.lower():
            return os.path.join(folder_path, file)
    return None

def process_folder(folder_path: str) -> Dict[str, Union[str, float]]:
    """
    Procesa una carpeta para extraer datos del Excel de cotización
    
    Args:
        folder_path: Ruta a la carpeta que contiene el Excel
    
    Returns:
        Diccionario con todos los datos extraídos estructurados
    """
    excel_file = find_excel_file(folder_path)
    if not excel_file:
        raise FileNotFoundError("No se encontró archivo Excel de cotización en la carpeta")
    
    file_name = os.path.basename(excel_file)
    return extract_excel_data(excel_file, file_name)

# Ejemplo de uso
if __name__ == "__main__":
    folder_path = input("Ingrese la ruta de la carpeta con el Excel: ")
    
    try:
        data = process_folder(folder_path)
        print("\nDatos extraídos:")
        for key, value in data.items():
            if key != "contenido_completo":  # No imprimir el contenido completo por brevedad
                print(f"{key.upper():<20}: {value}")
        
        print("\n(El contenido completo del Excel está disponible en la clave 'contenido_completo')")
    except Exception as e:
        logger.error(f"Error al procesar la carpeta: {str(e)}")